import os
from openpyxl import load_workbook
import configparser
import string
from operator import attrgetter

config_name = "config.ini"


class Record(object):

    def __init__(self, name, id, price, count):
        self.name = name
        self.price = price
        self.count = count
        self.id = id

    def __repr__(self):
        return repr((self.name, self.id, self.price, self.count))

    def record_id(self):
        return '{0}_{1}'.format(self.id, self.price)


class Excelfile(object):

    def __init__(self, file_name):
        self.file_name = file_name
        # Creale clear variables
        self.list = []
        self.header = {}
        # File params
        self.input_file = None
        self.output_file = None
        self.data_start = None
        # Col params
        self.col_name = None
        self.col_id = None
        self.col_measure = None
        self.col_price = None
        self.col_count = None
        self.col_total = None

        self.read_config()

    def read_xls_file(self):
        dict = {}
        file_path = '{0}/{1}'.format(os.getcwd(), self.input_file)
        if os.path.exists(file_path):
            wb = load_workbook(filename=file_path)
            ws = wb.active

            for i in range(self.data_start, ws.max_row - self.data_start):

                name = ws.cell(row=i, column=self.col_name).value
                if name == None:
                    name = ''
                rec = Record(
                    name,
                    ws.cell(row=i, column=self.col_id).value,
                    ws.cell(row=i, column=self.col_price).value,
                    ws.cell(row=i, column=self.col_count).value)

                rec2 = dict.get(rec.record_id())
                if rec2 != None:
                    rec.count += rec2.count
                    dict.pop(rec2.record_id())

                dict[rec.record_id()] = rec

            for value in dict.values():
                self.list.append(value)

            self.list.sort(key=attrgetter('name'))

    def write_xls_file(self):
        self.prepare_file_for_write()
        file_path = '{0}/{1}'.format(os.getcwd(), self.output_file)
        if os.path.exists(file_path):
            wb = load_workbook(filename=file_path)
            ws = wb.active

            i = self.data_start
            for item in self.list:
                ws.cell(row=i, column=self.col_name).value = item.name
                ws.cell(row=i, column=self.col_id).value = item.id
                ws.cell(row=i, column=self.col_measure).value = u'шт.'
                ws.cell(row=i, column=self.col_price).value = item.price
                ws.cell(row=i, column=self.col_count).value = item.count
                ws.cell(row=i, column=self.col_total).value = item.count * item.price

                i += 1

        wb.save(file_path)

    def read_config(self):
        if self.file_name.strip() != '':
            conf = configparser.ConfigParser()
            conf.read(self.file_name.strip())
            self.input_file = conf.get('FILE', 'InputFileName', fallback='input.xls')
            self.output_file = conf.get('FILE', 'OutputFileName', fallback='output.xls')
            # Start from 0, then title_height + 1
            self.data_start = int(conf.get('FILE', 'HeaderRowCount', fallback='1')) + 1
            # Read row params file
            self.col_name = int(conf.get('ROWS', 'ColName', fallback='2'))
            self.col_id = int(conf.get('ROWS', 'ColId', fallback='3'))
            self.col_measure = int(conf.get('ROWS', 'ColMeasure', fallback='4'))
            self.col_price = int(conf.get('ROWS', 'ColPrice', fallback='5'))
            self.col_count = int(conf.get('ROWS', 'ColCount', fallback='6'))
            self.col_total = int(conf.get('ROWS', 'ColTotal', fallback='7'))

    def prepare_file_for_write(self):
        # Copy file to destination path
        # Remove all rows except the title
        if os.path.exists('{0}/{1}'.format(os.getcwd(), self.input_file)):
            wb = load_workbook(filename='{0}/{1}'.format(os.getcwd(), self.input_file))
            ws = wb.active

            row = '{0}{1}'.format(string.ascii_uppercase[0], self.data_start)
            col = '{0}{1}'.format(string.ascii_uppercase[ws.max_column], ws.max_row)

            for rows in ws[row:col]:
                for i in range(0, ws.max_column):
                    rows[i].value = None

        wb.save('{0}/{1}'.format(os.getcwd(), self.output_file))


if __name__ == "__main__":
    xls = Excelfile(config_name)
    xls.read_xls_file()
    xls.write_xls_file()
